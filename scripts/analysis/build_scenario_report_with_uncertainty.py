# viz/build_scenario_report_with_uncertainty.py
"""
Build scenario comparison report WITH UNCERTAINTY QUANTIFICATION.

For each metric, shows:
- Mean (central estimate)
- Std Dev (absolute uncertainty)
- CV% (relative uncertainty)
- 5th/95th percentile (90% confidence interval)

This provides full uncertainty characterization from Monte Carlo iterations.
"""
import sys
import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# Add repo to path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

# Try to load project config for defaults (optional)
try:
    from fl_risk_model.config import cfg as _cfg
except Exception:
    _cfg = None


def _safe(x, default=np.nan):
    try:
        return float(x)
    except Exception:
        return default


def _safe_pct(numer, denom):
    numer = 0.0 if pd.isna(numer) else float(numer)
    denom = 0.0 if pd.isna(denom) else float(denom)
    if denom <= 0:
        return np.nan
    return 100.0 * (numer / denom)


def _autosize(ws, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_width, max(12, max_len + 2))


def compute_statistics(iter_df, scenario_name, metric_extractor):
    """
    Compute statistics for a metric across iterations.
    
    Parameters:
    -----------
    iter_df : DataFrame with iteration-level results
    scenario_name : str
    metric_extractor : callable that takes a row dict and returns metric value
    
    Returns:
    --------
    dict with keys: mean, std, cv_pct, p5, p95, median
    """
    scen_rows = iter_df[iter_df['scenario'] == scenario_name]
    if scen_rows.empty:
        return {
            'mean': np.nan, 'std': np.nan, 'cv_pct': np.nan,
            'p5': np.nan, 'p95': np.nan, 'median': np.nan
        }
    
    # Extract metric values
    values = []
    for _, row in scen_rows.iterrows():
        try:
            val = metric_extractor(row.to_dict())
            if not pd.isna(val):
                values.append(float(val))
        except Exception:
            continue
    
    if not values:
        return {
            'mean': np.nan, 'std': np.nan, 'cv_pct': np.nan,
            'p5': np.nan, 'p95': np.nan, 'median': np.nan
        }
    
    arr = np.array(values)
    mean_val = np.mean(arr)
    std_val = np.std(arr, ddof=1) if len(arr) > 1 else 0.0
    cv_val = (std_val / mean_val * 100.0) if mean_val != 0 else np.nan
    
    return {
        'mean': mean_val,
        'std': std_val,
        'cv_pct': cv_val,
        'p5': np.percentile(arr, 5),
        'p95': np.percentile(arr, 95),
        'median': np.median(arr)
    }


def build_report_with_uncertainty(iterations_csv, out_xlsx, fhcf_cap, nfip_fl_premium_base):
    """Build scenario report with full uncertainty quantification."""
    
    # Load iteration-level data
    iter_df = pd.read_csv(iterations_csv)
    if "scenario" not in iter_df.columns:
        raise ValueError("iterations CSV must contain a 'scenario' column.")
    
    # Map internal scenario IDs -> display names
    scen_map = {
        "great_miami": "Great Miami",
        "andrew": "Andrew",
        "double_gm": "Double Great Miami",
        "gm_plus_andrew": "Great Miami and Andrew",  # Legacy
        "andrew_then_gm": "Andrew then Great Miami",
        "gm_then_andrew": "Great Miami then Andrew",
        "lake_okeechobee": "Lake Okeechobee",
        "irma": "Irma",
        "double_irma": "Double Irma",
    }
    
    # Scenario display order
    col_order = [
        "Great Miami",
        "Andrew",
        "Double Great Miami",
        "Andrew then Great Miami",
        "Great Miami then Andrew",
        "Lake Okeechobee",
        "Irma",
        "Double Irma",
    ]
    
    # Define metrics (same as original report) with section headers
    # Each metric is (label, extractor_function) or (SECTION, section_name)
    # Note: iterations.csv uses _usd suffix, not _mean
    def _get(row, col, default=0.0):
        return row.get(col, default) if col in row else default
    
    # Calculate total damage on the fly
    def _total_damage(r):
        return _get(r, "wind_total_usd") + _get(r, "water_total_usd")
    
    SECTION = "SECTION"  # Special marker for section headers
    
    # Layout matches original report structure
    layout = [
        (SECTION, "Scenario & Total Impact"),
        ("Total loss (USD)", lambda r: _total_damage(r)),
        ("Wind loss gross (USD)", lambda r: _get(r, "wind_total_usd", np.nan)),
        ("Wind loss gross (%)", lambda r: _safe_pct(_get(r, "wind_total_usd"), _total_damage(r))),
        ("Flood loss gross (USD)", lambda r: _get(r, "water_total_usd", np.nan)),
        ("Flood loss gross (%)", lambda r: _safe_pct(_get(r, "water_total_usd"), _total_damage(r))),
        ("Citizens loss gross (USD)", lambda r: _get(r, "wind_insured_citizens_usd", np.nan)),
        ("Citizens loss gross (%)", lambda r: _safe_pct(_get(r, "wind_insured_citizens_usd"), _total_damage(r))),
        ("Un/underinsured wind (USD)", lambda r: _get(r, "wind_uninsured_usd") + _get(r, "wind_underinsured_usd")),
        ("Un/underinsured wind (%)", lambda r: _safe_pct(_get(r, "wind_uninsured_usd") + _get(r, "wind_underinsured_usd"), _total_damage(r))),
        ("Un/underinsured flood (USD)", lambda r: _get(r, "flood_un_derinsured_usd", np.nan)),
        ("Un/underinsured flood (%)", lambda r: _safe_pct(_get(r, "flood_un_derinsured_usd"), _total_damage(r))),
        ("Total household burden (USD)", lambda r: (_get(r, "wind_uninsured_usd") + _get(r, "wind_underinsured_usd") + _get(r, "flood_un_derinsured_usd"))),
        ("Total household burden (%)", lambda r: _safe_pct((_get(r, "wind_uninsured_usd") + _get(r, "wind_underinsured_usd") + _get(r, "flood_un_derinsured_usd")), _total_damage(r))),
        
        (SECTION, "Institutional Stress"),
        ("FHCF recovery (private) (USD)", lambda r: _get(r, "fhcf_recovery_private_usd", np.nan)),
        ("FHCF recovery (Citizens) (USD)", lambda r: _get(r, "fhcf_recovery_citizens_usd", np.nan)),
        # FHCF cap binds: for individual iteration this is boolean, for mean it's the % of iterations where cap binds
        ("FHCF cap binds (%)", lambda r: (100.0 if _get(r, "fhcf_cap_binding", False) else 0.0)),
        ("FHCF shortfall (USD)", lambda r: _get(r, "fhcf_shortfall_usd", np.nan)),
        ("Cat bond payout (USD)", lambda r: _get(r, "catbond_payout_usd", np.nan)),
        ("Citizens Tier 1 (USD)", lambda r: _get(r, "citizens_tier1_usd", np.nan)),
        ("Citizens Tier 2 (USD)", lambda r: _get(r, "citizens_tier2_usd", np.nan)),
        ("Citizens residual (USD)", lambda r: _get(r, "citizens_residual_deficit_usd", np.nan)),
        ("FIGA collections (USD)", lambda r: _get(r, "figa_collected_usd", np.nan)),
        ("FIGA residual (USD)", lambda r: _get(r, "figa_residual_deficit_usd", np.nan)),
        ("NFIP payouts (total) (USD)", lambda r: _get(r, "nfip_claims_paid_usd", np.nan)),
        ("NFIP national pool used (USD)", lambda r: _get(r, "nfip_pool_used_usd", np.nan)),
        ("NFIP Treasury borrowing (USD)", lambda r: _get(r, "nfip_borrowed_usd", np.nan)),
        ("Total public burden (USD)", lambda r: (_get(r, "citizens_residual_deficit_usd") + _get(r, "figa_residual_deficit_usd") + _get(r, "nfip_borrowed_usd") + _get(r, "fhcf_shortfall_usd"))),
        
        (SECTION, "Capital & Default Outcomes"),
        ("Private defaults pre-group (#)", lambda r: _get(r, "defaults_pre", np.nan)),
        ("Private defaults post-group (#)", lambda r: _get(r, "defaults_post", np.nan)),
        # defaulted_premium_base_pct is already 0-100, don't multiply
        ("Share of defaulted premium base (%)", lambda r: _get(r, "defaulted_premium_base_pct", np.nan)),
        ("Largest single-entity deficit (USD)", lambda r: _get(r, "largest_entity_deficit_usd", np.nan)),
        ("Group shortfall total (USD)", lambda r: _get(r, "diag_group_shortfall_total_usd", np.nan)),
        ("Groups with shortfall (#)", lambda r: _get(r, "diag_groups_with_shortfall", np.nan)),
        # diag_groups_fully_funded_share is 0-1, multiply by 100
        ("Groups fully funded share (%)", lambda r: _get(r, "diag_groups_fully_funded_share", np.nan) * 100.0 if "diag_groups_fully_funded_share" in r else np.nan),
        
        (SECTION, "Ratios / Stress Indicators"),
        ("FHCF utilization ratio (%)", lambda r: _safe_pct(_get(r, "fhcf_recovery_private_usd") + _get(r, "fhcf_recovery_citizens_usd"), fhcf_cap) if fhcf_cap else np.nan),
        ("Citizens assessment stress ratio (%)", lambda r: _safe_pct(_get(r, "citizens_residual_deficit_usd"), _get(r, "citizens_tier1_capacity_usd") + _get(r, "citizens_tier2_capacity_usd"))),
        ("FIGA stress ratio (%)", lambda r: _safe_pct(_get(r, "figa_residual_deficit_usd"), _get(r, "figa_residual_deficit_usd") + _get(r, "figa_collected_usd"))),
        ("NFIP Florida stress ratio (%)", lambda r: _safe_pct(_get(r, "nfip_claims_paid_usd"), _get(r, "nfip_fl_premium_base_usd")) if _get(r, "nfip_fl_premium_base_usd") else np.nan),
    ]
    
    # Extract just the metrics (non-section rows) for statistics computation
    metric_definitions = {label: extractor for label, extractor in layout if label != SECTION}
    
    # Build statistics for all scenarios and metrics
    print("Computing statistics for all scenarios and metrics...")
    stats_data = {}
    for scen_display in col_order:
        # Find internal scenario name
        scen_internal = None
        for k, v in scen_map.items():
            if v == scen_display:
                scen_internal = k
                break
        
        if scen_internal is None:
            continue
        
        stats_data[scen_display] = {}
        for metric_name, extractor in metric_definitions.items():
            stats = compute_statistics(iter_df, scen_internal, extractor)
            stats_data[scen_display][metric_name] = stats
    
    print(f"Computed statistics for {len(col_order)} scenarios and {len(metric_definitions)} metrics")
    
    # Build Excel workbook
    wb = Workbook()
    
    # Formatting styles
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    
    # Helper to build a sheet with section headers
    def build_sheet(wb, sheet_name, stat_key):
        ws = wb.create_sheet(sheet_name) if sheet_name != "Mean Values" else wb.active
        ws.title = sheet_name
        
        # Header row
        header = ["Metric"] + col_order
        ws.append(header)
        for j in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=j)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows with section headers
        for label, extractor in layout:
            if label == SECTION:
                # Section header row
                ws.append([extractor] + [""] * len(col_order))
                row_idx = ws.max_row
                for j in range(1, len(header) + 1):
                    cell = ws.cell(row=row_idx, column=j)
                    cell.font = section_font
                    cell.fill = section_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                # Metric data row
                row_data = [label]
                for scen in col_order:
                    val = stats_data.get(scen, {}).get(label, {}).get(stat_key, np.nan)
                    # Convert percentage values to decimal for Excel percentage format
                    # Percentages are stored as 0-100, but Excel %format expects 0-1
                    if not pd.isna(val) and "(%)" in label:
                        val = val / 100.0
                    row_data.append(val if not pd.isna(val) else "")
                ws.append(row_data)
        
        return ws
    
    # === SHEET 1: Mean Values ===
    ws_mean = build_sheet(wb, "Mean Values", 'mean')
    
    # === SHEET 2: Standard Deviation ===
    ws_std = build_sheet(wb, "Standard Deviation", 'std')
    
    # === SHEET 3: Coefficient of Variation (%) ===
    ws_cv = build_sheet(wb, "CV (%)", 'cv_pct')
    
    # === SHEET 4: 5th Percentile ===
    ws_p5 = build_sheet(wb, "5th Percentile", 'p5')
    
    # === SHEET 5: 95th Percentile ===
    ws_p95 = build_sheet(wb, "95th Percentile", 'p95')
    
    # === SHEET 6: Median ===
    ws_median = build_sheet(wb, "Median", 'median')
    
    # Format all sheets
    fmt_currency = '"$"#,##0'
    fmt_count = '#,##0'
    fmt_number = '#,##0.0'
    fmt_percent = '0.0%'
    
    def get_format(label):
        """Determine number format based on metric label."""
        if label == SECTION:
            return None
        if re.search(r"\(%\)", label):
            return fmt_percent
        if re.search(r"\(#\)", label):
            return fmt_count
        if re.search(r"\(USD\)", label, flags=re.I):
            return fmt_currency
        return fmt_number
    
    for ws in [ws_mean, ws_std, ws_cv, ws_p5, ws_p95, ws_median]:
        for row_idx in range(2, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            
            # Skip section header rows
            if label and any(label == section_name for _, section_name in layout if _ == SECTION):
                continue
            
            fmt = get_format(label) if label else fmt_number
            
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Autosize
        _autosize(ws)
    
    # === SHEET 7: Metric Dictionary ===
    desc_ws = wb.create_sheet("Metric dictionary")
    desc_ws.append(["Metric", "Definition"])
    desc_ws["A1"].font = Font(bold=True)
    desc_ws["B1"].font = Font(bold=True)
    
    metric_descriptions = [
        ("Total loss (USD)", "Aggregate economic loss across all hazards and payers."),
        ("Wind loss gross (USD)", "Gross hurricane wind losses before recoveries."),
        ("Wind loss gross (%)", "Share of total loss that is wind (gross)."),
        ("Flood loss gross (USD)", "Gross flood losses before NFIP recoveries."),
        ("Flood loss gross (%)", "Share of total loss that is flood (gross)."),
        ("Citizens loss gross (USD)", "Gross wind losses in Citizens' portfolio."),
        ("Citizens loss gross (%)", "Share of total loss in Citizens' gross wind."),
        ("Un/underinsured wind (USD)", "Wind damages not covered by insurance or above policy limits."),
        ("Un/underinsured wind (%)", "Share of total loss that is wind uninsured/underinsured."),
        ("Un/underinsured flood (USD)", "Flood damages not covered by NFIP or above policy caps."),
        ("Un/underinsured flood (%)", "Share of total loss that is flood uninsured/underinsured."),
        ("Total household burden (USD)", "Un/underinsured wind + flood (policyholder burden)."),
        ("Total household burden (%)", "Share of total loss borne by households (un/underinsured)."),
        ("FHCF recovery (private) (USD)", "Total FHCF reimbursement to private insurers."),
        ("FHCF recovery (Citizens) (USD)", "FHCF reimbursement to Citizens."),
        ("FHCF cap binds (%)", "Percentage of iterations where the statewide $17B cap is exceeded."),
        ("FHCF shortfall (USD)", "Recovery lost due to the cap binding."),
        ("Cat bond payout (USD)", "Aggregate payout from triggered catastrophe bonds."),
        ("Citizens Tier 1 (USD)", "Realized Tier 1 assessment on Citizens policyholders."),
        ("Citizens Tier 2 (USD)", "Realized Tier 2 assessment on all Florida policyholders."),
        ("Citizens residual (USD)", "Citizens remaining deficit after assessments."),
        ("FIGA collections (USD)", "Assessments collected from solvent insurers."),
        ("FIGA residual (USD)", "Unfunded deficit after FIGA max capacity."),
        ("NFIP payouts (total) (USD)", "NFIP claims paid attributable to Florida."),
        ("NFIP national pool used (USD)", "Portion taken from the national NFIP reserve/pool."),
        ("NFIP Treasury borrowing (USD)", "Shortfall financed via U.S. Treasury borrowing."),
        ("Total public burden (USD)", "FHCF shortfall + Citizens residual + FIGA residual + NFIP borrowing."),
        ("Private defaults pre-group (#)", "Insurer entities insolvent before group support."),
        ("Private defaults post-group (#)", "Insurer entities insolvent after group support."),
        ("Share of defaulted premium base (%)", "Statewide premium share held by defaulted insurers."),
        ("Largest single-entity deficit (USD)", "Maximum capital shortfall of any one insurer."),
        ("Group shortfall total (USD)", "Total group-level shortfall remaining after support."),
        ("Groups with shortfall (#)", "Number of groups with a remaining shortfall."),
        ("Groups fully funded share (%)", "Share of groups that are fully funded after support."),
        ("FHCF utilization ratio (%)", "Share of the $17B cap utilized."),
        ("Citizens assessment stress ratio (%)", "Citizens residual ÷ (Tier 1 + Tier 2 max capacity)."),
        ("FIGA stress ratio (%)", "Default deficit ÷ FIGA's max assessment capacity (approx)."),
        ("NFIP Florida stress ratio (%)", "Florida NFIP claims paid ÷ Florida NFIP premium base."),
    ]
    
    for m, d in metric_descriptions:
        desc_ws.append([m, d])
    
    _autosize(desc_ws, max_width=100)
    
    # Save
    wb.save(out_xlsx)
    print(f"✅ Saved uncertainty report: {out_xlsx}")
    print(f"   - 7 sheets: Mean, Std Dev, CV%, p5, p95, Median, Metric dictionary")
    print(f"   - {len(metric_definitions)} metrics × {len(col_order)} scenarios")


def main():
    default_fhcf_cap = getattr(_cfg, "FHCF_SEASON_CAP", 17_000_000_000.0)
    
    ap = argparse.ArgumentParser(
        description="Build scenario report with uncertainty quantification from MC iterations"
    )
    ap.add_argument("--iterations", required=True, help="Path to iterations.csv (iteration-level results)")
    ap.add_argument("--out", required=True, help="Output Excel path")
    ap.add_argument("--fhcf_cap", type=float, default=default_fhcf_cap,
                    help="FHCF statewide cap (USD)")
    ap.add_argument("--nfip_fl_premium_base", type=float, default=None,
                    help="Override Florida NFIP premium base (USD)")
    args = ap.parse_args()
    
    build_report_with_uncertainty(args.iterations, args.out, args.fhcf_cap, args.nfip_fl_premium_base)


if __name__ == "__main__":
    main()
